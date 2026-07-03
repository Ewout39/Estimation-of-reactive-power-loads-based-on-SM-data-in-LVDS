import os
import zipfile
import pandas as pd
import shutil
import numpy as np
import shutil
from openpyxl import load_workbook
from sklearn.metrics.cluster import fowlkes_mallows_score
from sklearn.preprocessing import MinMaxScaler
from kneed import KneeLocator
from collections import Counter
from scipy.cluster.hierarchy import dendrogram
from sklearn.cluster import KMeans
import matplotlib.pyplot as plt
from tslearn.clustering import TimeSeriesKMeans
from sklearn.metrics import silhouette_score
from scipy.spatial.distance import pdist
from scipy.cluster.hierarchy import linkage, fcluster
from scipy.spatial.distance import squareform


def loading_cleaned_files(path_folder, dict_name, dataset_chosen):
    """Loads cleaned data files from a zip archive and processes them based on the chosen dataset.
    
    Args:
        path_folder (Path): Path to the folder containing the zip archive.
        dict_name (str): Name of the zip archive file.
        dataset_chosen (str): Name of the dataset chosen ("Slovakia", "Germany", or "USA").
    Returns:
        dict: A dictionary containing the cleaned dataframes for each household.
    """
    if dataset_chosen == "Slovakia":
        col_name = "datetime"
        area_code = 'Europe/Bratislava'
    elif dataset_chosen == "Germany":
        col_name = "index"
        area_code = 'Europe/Berlin'
    elif dataset_chosen == "USA":
        col_name = "DateTimeUTC"
        area_code = 'America/New_York'

    temp_dir = path_folder / "extracted_dataframes"
    temp_dir.mkdir(exist_ok=True)

    powerdf_clean = {}

    zip_path = path_folder / dict_name

    with zipfile.ZipFile(zip_path, 'r') as zipf:
        zipf.extractall(temp_dir)

    for file in temp_dir.iterdir():
        if file.suffix == ".csv":
            try:
                key = int(file.stem)
            except ValueError:
                continue

            df = pd.read_csv(file, index_col=col_name, parse_dates=[col_name])

            
            if not isinstance(df.index, pd.DatetimeIndex):
                        df.index = pd.to_datetime(df.index, errors='coerce')


            if df.index.tz is None:
                 df.index = df.index.tz_localize(area_code)
            else:
                df.index = df.index.tz_convert(area_code)

            powerdf_clean[key] = df

    if temp_dir.exists():
            shutil.rmtree(temp_dir, onerror=on_error)

    powerdf_clean = {k: powerdf_clean[k] for k in sorted(powerdf_clean)}

    houses = []
    for i in powerdf_clean.keys():
        houses.append(int(powerdf_clean[i]['Nr'].iloc[1]))
    
    return powerdf_clean, houses

def on_error(function, path, exc_info):
    """Prints error information when a file or directory cannot be removed.
    
    Args:
        path (str): Path of the file that couldn't be removed.
        exc_info (tuple): Exception information.
    """
    print(f"Error removing {path}: {exc_info}")

def basecase(power_dict, MAEVsRMSE):
    Estimation_error = []
    PF_range = [round(x * 0.01, 2) for x in range(81, 101)]
    for PF in PF_range:
        error_inner = 0
        for house, power_data in power_dict.items():
            P_values = power_data['P'].values
            Q_values = power_data['Q'].values
            Q_est_inner = np.sqrt((P_values/PF)**2-(P_values)**2)
            if MAEVsRMSE == "MAE":
                error_inner += np.sum(abs(Q_values-Q_est_inner))
            else:
                error_inner += np.sum((Q_values - Q_est_inner)**2)
        if MAEVsRMSE == "MAE":
            error_tot = error_inner/(len(power_dict)*len(power_data))
        else:
            error_tot = np.sqrt(error_inner/(len(power_dict)*len(power_data)))
        Estimation_error.append(float(round(error_tot, 5)))
    basecase = min(Estimation_error)
    best_index = Estimation_error.index(basecase)
    best_PF = PF_range[best_index]
    print(Estimation_error)
    return best_PF, basecase

def optimal_error(dataset_timed, dataset, MAEVsRMSE):
    """Calculates the optimal error and corresponding power factor (PF) for each household in the dataset.

    Args:
        dataset_timed (dict): Dictionary containing power dataframes for each household with time-indexed data.
        dataset (dict): Dictionary containing power dataframes for each household.
        MAEVsRMSE (str): The error metric to use ("MAE" or "RMSE").

    Returns:
        float: The total error across all households.
        list: A list of tuples containing the household number and its corresponding optimal PF.
    """
    error_tot = 0
    PF_list = []
    PF_range = [round(x * 0.01, 2) for x in range(60, 101)]
    for (house, power_data), (house2, power_data2) in zip(dataset_timed.items(), dataset.items()):
        P_values = power_data['P'].values
        Q_values = power_data['Q'].values
        Nr = power_data2['Nr'].iloc[0]
        error_inner_min = 1000000000000
        for PF in PF_range:
            error_inner = 0
            Q_est_inner = np.sqrt((P_values/PF)**2-(P_values)**2)
            if MAEVsRMSE == "RMSE":
                error_inner += np.sum((Q_values - Q_est_inner)**2)
            else:
                error_inner = np.sum(np.abs(Q_values - Q_est_inner))
            if error_inner < error_inner_min:
                error_inner_min = error_inner
                optimal_PF = PF
        PF_list.append((Nr, optimal_PF))

    indx = 0
    for house, power_data in dataset.items():
        P_values = power_data['P'].values
        Q_values = power_data['Q'].values
        PF = PF_list[indx][1]
        Q_est_inner = np.sqrt((P_values/PF)**2-(P_values)**2)
        if MAEVsRMSE == "RMSE":
            error_inner = np.sum((Q_values - Q_est_inner)**2)
        else:
            error_inner = np.sum(np.abs(Q_values - Q_est_inner))
        error_tot += error_inner
        indx += 1

    return error_tot, PF_list

def adapting_data_resolution(power_dict, target_resolution, dataset_chosen):
    """Adapts the data resolution of the power dictionary to the target resolution.
    
    Args:
        power_dict (dict): Dictionary containing power dataframes for each household.
        target_resolution (str): Target resolution for resampling (e.g., '15min', '1H').
    Returns:
        dict: A new dictionary with resampled dataframes at the target resolution.
        int: Length of a week in the target resolution.
        int: Length of a month (first type) in the target resolution.
        int: Length of a month (second type) in the target resolution.
        int: Length of the summer season in the target resolution.
        int: Length of the winter season in the target resolution.
        int: Length of the fall season in the target resolution.
        int: Length of the spring season in the target resolution.
    """

    year = power_dict[list(power_dict.keys())[0]].index.year[100]
    
    power_dict_timed = {}
    for house, power_data in power_dict.items():
        if target_resolution == "15_min":
            power_data_timed = power_data
            week_length = 7*24*4
            month_length1 = 30*24*4
            month_length2 = 31*24*4
            Summer_length = (31+31+30)*24*4
            if is_leap_year(year):
                if dataset_chosen == "Slovakia":
                    Winter_length = (31+31+29+29)*24*4 + 5
                else:
                    Winter_length = (31+31+29)*24*4
            else:
                Winter_length = (31+31+28)*24*4
            Fall_length = (30+31+30)*24*4 + 4
            Spring_length = (31+31+30)*24*4-4
        elif target_resolution == "hourly":
            power_data_timed = power_data.resample('h').mean()
            week_length = 7*24
            month_length1 = 30*24
            month_length2 = 31*24
            Summer_length = (31+31+30)*24
            if is_leap_year(year):
                if dataset_chosen == "Slovakia":
                    Winter_length = (31+31+29+29)*24 + 2
                else:
                    Winter_length = (31+31+29)*24
            else:
                Winter_length = (31+31+28)*24
            Fall_length = (30+31+30)*24 + 1
            Spring_length = (31+31+30)*24-1
        else:
            power_data_timed = power_data.resample('d').mean()
            week_length = 7
            month_length1 = 30
            month_length2 = 31
            Summer_length = (31+31+30)
            if is_leap_year(year):
                if dataset_chosen == "Slovakia":
                    Winter_length = (31+31+29+29) + 1
                else:
                    Winter_length = (31+31+29)
            else:
                Winter_length = (31+31+28)
            Fall_length = (30+31+30)
            Spring_length = (31+31+30)
        power_dict_timed[house] = power_data_timed
    return power_dict_timed, month_length2, Summer_length, Winter_length, Fall_length, Spring_length

def is_leap_year(year):
    """Checks if a given year is a leap year.
    
    Args:
        year (int): The year to check.
    Returns:
        bool: True if the year is a leap year, False otherwise.
    """
    return (year % 4 == 0 and year % 100 != 0) or (year % 400 == 0)

def creating_season_data(dataset_timed, dataset, Summer_length, Winter_length, Fall_length, Spring_length, dataset_chosen, variable, Clustering_method):
    monthly_data = {}
    for i in range(1, len(dataset_timed)+1):
        if Clustering_method == "Unimportant":
            monthly_data[i] = [month[variable].reset_index(drop=True) for _, month in dataset_timed[i].groupby(pd.Grouper(freq='ME'))]
        else:
            monthly_data[i] = [month[variable].reset_index(drop=True) for _, month in dataset_timed[i].groupby(pd.Grouper(freq='ME'))]

    if Clustering_method == "PF_time":
        Summer_data_timed = pd.DataFrame(np.zeros((Summer_length, len(dataset_timed))), columns = range(1, len(dataset_timed)+1))
        Winter_data_timed = pd.DataFrame(np.zeros((Winter_length, len(dataset_timed))), columns = range(1, len(dataset_timed)+1))
        Fall_data_timed = pd.DataFrame(np.zeros((Fall_length, len(dataset_timed))), columns = range(1, len(dataset_timed)+1))
        Spring_data_timed = pd.DataFrame(np.zeros((Spring_length, len(dataset_timed))), columns = range(1, len(dataset_timed)+1))
    else:
        Summer_data_timed = pd.DataFrame(np.zeros((Summer_length, len(dataset_timed))))
        Winter_data_timed = pd.DataFrame(np.zeros((Winter_length, len(dataset_timed))))
        Fall_data_timed = pd.DataFrame(np.zeros((Fall_length, len(dataset_timed))))
        Spring_data_timed = pd.DataFrame(np.zeros((Spring_length, len(dataset_timed))))

    initial_index = dataset_timed[1].index
    Winter_index = initial_index[initial_index.month.isin([12, 1, 2])]
    Summer_index = initial_index[initial_index.month.isin([6, 7, 8])]
    Spring_index = initial_index[initial_index.month.isin([3, 4, 5])]
    Fall_index = initial_index[initial_index.month.isin([9, 10, 11])]

    if dataset_chosen == "Slovakia":
        winter_months = {0, 1, 11, 12}
    else:
        winter_months = {0, 1, 11}

    for i in range(1, len(dataset_timed)+1):
        row_winter = 0
        row_spring = 0
        row_summer = 0
        row_fall = 0
        for j, month in enumerate(monthly_data[i]):
            if j in winter_months:
                Winter_data_timed.iloc[row_winter:row_winter+len(month),i-1] = month
                row_winter += len(month)
            elif j in {2, 3, 4}:
                Spring_data_timed.iloc[row_spring:row_spring+len(month),i-1] = month
                row_spring += len(month)
            elif j in {5, 6, 7}:
                Summer_data_timed.iloc[row_summer:row_summer+len(month), i-1] = month
                row_summer += len(month)
            else:
                Fall_data_timed.iloc[row_fall:row_fall+len(month), i-1] = month
                row_fall += len(month)

    Winter_data_timed.index = Winter_index
    Summer_data_timed.index = Summer_index
    Spring_data_timed.index = Spring_index
    Fall_data_timed.index = Fall_index

    Summer_data = {}
    Fall_data = {}
    Winter_data = {}
    Spring_data = {}
    for ind, power_data in dataset.items():
        summer_data = power_data[power_data.index.month.isin([6, 7, 8])]
        fall_data = power_data[power_data.index.month.isin([9, 10, 11])]
        winter_data = power_data[power_data.index.month.isin([12, 1, 2])]
        spring_data = power_data[power_data.index.month.isin([3, 4, 5])]
        Summer_data[ind] = summer_data
        Fall_data[ind] = fall_data
        Winter_data[ind] = winter_data
        Spring_data[ind] = spring_data

    return Summer_data_timed, Fall_data_timed, Winter_data_timed, Spring_data_timed, Summer_data, Fall_data, Winter_data, Spring_data

def writing_results_training(input_scenario, path_folder, dataset_chosen, clustering_method, resolution, input_period, MAEVsRMSE, improvement, error, silhouette, Nr_clusters, random_state, correlating_features, Nr_PCA_components):
    """Writes the clustering results and hyperparameters to an Excel file.

    Args:
        input_scenario (str): The input scenario for clustering.
        path_folder (Path): Path to the folder containing the Excel file.
        dataset_chosen (str): Name of the dataset chosen ("Slovakia", "Germany", or "USA").
        clustering_method (str): The clustering method used.
        resolution (str): The data resolution used.
        input_period (str): The input period for clustering.
        MAEVsRMSE (str): The error metric used ("MAE" or "RMSE").
        improvement (float): The improvement value to write.
        error (float): The error value to write.
        silhouette (float): The silhouette score to write.
        Nr_clusters (list): List of number of clusters used in clustering.
        random_state (list): List of random states used in clustering.
        correlating_features (list): List of correlating features used in clustering.
        Nr_PCA_components (list): List of number of PCA components used in clustering.
    """
    template_path = path_folder / "Clustering_overview.xlsx"
    file_path = path_folder / dataset_chosen / f"Clustering_overview_{dataset_chosen}.xlsx"

    if not os.path.exists(file_path):
        shutil.copy(template_path, file_path)

    wb = load_workbook(file_path)
    ws_hyperparameters = wb['Hyperparameters']
    ws_results = wb[f'{clustering_method}_training_results']

    if input_scenario == "PF_time":
        column_improvement = "E"
        column_error = "N"
        column_hyperparameters_clusters = "C"
        column_hyperparameters_random_state = "J"
        column_hyperparameters_PCA = "C"
        if MAEVsRMSE == "RMSE":
            column_silhouette = "C"
    elif input_scenario == "PQ_Pearson":
        column_improvement = "F"
        column_error = "O"
        column_hyperparameters_clusters = "D"
        column_hyperparameters_random_state = "K"
        column_hyperparameters_PCA = "D"
        if MAEVsRMSE == "RMSE":
            column_silhouette = "D"
    elif input_scenario == "PQSPF_Pearson":
        column_improvement = "G"
        column_error = "P"
        column_hyperparameters_clusters = "E"
        column_hyperparameters_random_state = "L"
        column_hyperparameters_PCA = "E"
        if MAEVsRMSE == "RMSE":
            column_silhouette = "E"
    elif input_scenario == "PQ_PCA":
        column_improvement = "H"
        column_error = "Q"
        column_hyperparameters_clusters = "F"
        column_hyperparameters_random_state = "M"
        column_hyperparameters_PCA = "F"
        if MAEVsRMSE == "RMSE":
            column_silhouette = "F"
    elif input_scenario == "PQSPF_PCA":
        column_improvement = "I"
        column_error = "R"
        column_hyperparameters_clusters = "G"
        column_hyperparameters_random_state = "N"
        column_hyperparameters_PCA = "G"
        if MAEVsRMSE == "RMSE":
            column_silhouette = "G"

    if MAEVsRMSE == "MAE":
        if input_period == "year":
            if resolution == "15_min":
                row = 5
            elif resolution == "hourly":
                row = 6
            else:
                row = 7
        elif input_period == "month":
            if resolution == "15_min":
                row = 8
            elif resolution == "hourly":
                row = 9
            else:
                row = 10
        elif input_period == "seasonal":
            if resolution == "15_min":
                row = 11
            elif resolution == "hourly":
                row = 12
            else:
                row = 13
    else:
        if input_period == "year":
            if resolution == "15_min":
                row = 18
                row_KMeans = 5
                row_hierarchical = 18
                row_PCA = 31
                row_silhouette = 31
            elif resolution == "hourly":
                row = 19
                row_KMeans = 6
                row_hierarchical = 19
                row_PCA = 32
                row_silhouette = 32
            else:
                row = 20
                row_KMeans = 7
                row_hierarchical = 20
                row_PCA = 33
                row_silhouette = 33
        elif input_period == "month":
            if resolution == "15_min":
                row = 21
                row_KMeans = 8
                row_hierarchical = 21
                row_PCA = 34
                row_silhouette = 34
            elif resolution == "hourly":
                row = 22
                row_KMeans = 9
                row_hierarchical = 22
                row_PCA = 35
                row_silhouette = 35
            else:
                row = 23
                row_KMeans = 10
                row_hierarchical = 23
                row_PCA = 36
                row_silhouette = 36
        elif input_period == "seasonal":
            if resolution == "15_min":
                row = 24
                row_KMeans = 11
                row_hierarchical = 24
                row_PCA = 37
                row_silhouette = 37
            elif resolution == "hourly":
                row = 25
                row_KMeans = 12
                row_hierarchical = 25
                row_PCA = 38
                row_silhouette = 38
            else:
                row = 26
                row_KMeans = 13
                row_hierarchical = 26
                row_PCA = 39
                row_silhouette = 39

    ws_results[f"{column_improvement}{row}"].value = improvement
    ws_results[f"{column_error}{row}"].value = error
    if MAEVsRMSE == "RMSE":
        ws_results[f"{column_silhouette}{row_silhouette}"].value = silhouette
        if clustering_method == "KMeans":
            ws_hyperparameters[f"{column_hyperparameters_clusters}{row_KMeans}"].value = ", ".join(map(str, Nr_clusters))
            ws_hyperparameters[f"{column_hyperparameters_random_state}{row_KMeans}"].value = ", ".join(map(str, random_state))
        else:
            ws_hyperparameters[f"{column_hyperparameters_clusters}{row_hierarchical}"].value = ", ".join(map(str, Nr_clusters))
        if input_scenario in ["PQ_PCA", "PQSPF_PCA"]:
            ws_hyperparameters[f"{column_hyperparameters_PCA}{row_PCA}"].value = ", ".join(map(str, Nr_PCA_components))
        elif input_scenario in ["PQ_Pearson", "PQSPF_Pearson"]:
            ws_hyperparameters[f"{column_hyperparameters_PCA}{row_PCA}"].value =correlating_features
    wb.save(file_path)

def writing_results_testing(input_scenario, path_folder, dataset_chosen, clustering_method, resolution, input_period, MAEVsRMSE, improvement, error, silhouette):
    """Writes the clustering results from testing to an Excel file.

    Args:
        input_scenario (str): The input scenario for clustering.
        path_folder (Path): Path to the folder containing the Excel file.
        dataset_chosen (str): Name of the dataset chosen ("Slovakia", "Germany", or "USA").
        clustering_method (str): The clustering method used.
        resolution (str): The data resolution used.
        input_period (str): The input period for clustering.
        MAEVsRMSE (str): The error metric used ("MAE" or "RMSE").
        improvement (float): The improvement value to write.
        error (float): The error value to write.
        silhouette (float): The silhouette score to write.
    """
    template_path = path_folder / "Clustering_overview.xlsx"
    file_path = path_folder / dataset_chosen / f"Clustering_overview_{dataset_chosen}.xlsx"


    if not os.path.exists(file_path):
        shutil.copy(template_path, file_path)

    wb = wb = load_workbook(file_path)
    ws_results = wb[f'{clustering_method}_testing_results']

    if input_scenario == "PF_time":
        column_improvement = "E"
        column_error = "N"
        if MAEVsRMSE == "RMSE":
            column_silhouette = "C"
    elif input_scenario == "PQ_Pearson":
        column_improvement = "F"
        column_error = "O"
        if MAEVsRMSE == "RMSE":
            column_silhouette = "D"
    elif input_scenario == "PQSPF_Pearson":
        column_improvement = "G"
        column_error = "P"
        if MAEVsRMSE == "RMSE":
            column_silhouette = "E"
    elif input_scenario == "PQ_PCA":
        column_improvement = "H"
        column_error = "Q"
        if MAEVsRMSE == "RMSE":
            column_silhouette = "F"
    elif input_scenario == "PQSPF_PCA":
        column_improvement = "I"
        column_error = "R"
        if MAEVsRMSE == "RMSE":
            column_silhouette = "G"

    if MAEVsRMSE == "MAE":
        if input_period == "year":
            if resolution == "15_min":
                row = 18
            elif resolution == "hourly":
                row = 19
            else:
                row = 20
        elif input_period == "month":
            if resolution == "15_min":
                row = 21
            elif resolution == "hourly":
                row = 22
            else:
                row = 23
        elif input_period == "seasonal":
            if resolution == "15_min":
                row = 24
            elif resolution == "hourly":
                row = 25
            else:
                row = 26
    else:
        if input_period == "year":
            if resolution == "15_min":
                row = 5
                row_silhouette = 31
            elif resolution == "hourly":
                row = 6
                row_silhouette = 32
            else:
                row = 7
                row_silhouette = 33
        elif input_period == "month":
            if resolution == "15_min":
                row = 8
                row_silhouette = 34
            elif resolution == "hourly":
                row = 9
                row_silhouette = 35
            else:
                row = 10
                row_silhouette = 36
        elif input_period == "seasonal":
            if resolution == "15_min":
                row = 11
                row_silhouette = 37
            elif resolution == "hourly":
                row = 12
                row_silhouette = 38
            else:
                row = 13
                row_silhouette = 39

    ws_results[f"{column_improvement}{row}"].value = improvement
    ws_results[f"{column_error}{row}"].value = error
    if MAEVsRMSE == "RMSE":
        ws_results[f"{column_silhouette}{row_silhouette}"].value = silhouette
        
    wb.save(file_path)

def writing_results_benchmark_scenarios(path_folder, dataset_chosen, trainVsTest, resolution, input_period, MAEVsRMSE, error_basecase, error_tub, improvement_tub):
    """Writes the clustering results from testing to an Excel file.

    Args:
        path_folder (Path): Path to the folder containing the Excel file.
        dataset_chosen (str): Name of the dataset chosen ("Slovakia", "Germany", or "USA").
        trainVsTest (str): Whether to write training or testing results.
        resolution (str): The data resolution used.
        input_period (str): The input period for clustering.
        MAEVsRMSE (str): The error metric used ("MAE" or "RMSE").
        error_basecase (float): The error value for the base case.
        error_tub (float): The error value for the TUB method.
        improvement_tub (float): The improvement value for the TUB method.
    """
    template_path = path_folder / "Clustering_overview.xlsx"
    file_path = path_folder / dataset_chosen / f"Clustering_overview_{dataset_chosen}.xlsx"


    if not os.path.exists(file_path):
        shutil.copy(template_path, file_path)

    wb = wb = load_workbook(file_path)
    if trainVsTest == "Train":
        ws_KMeans = wb[f'KMeans_training_results']
        ws_hierarchical = wb[f'hierarchical_training_results']
    else:
        ws_KMeans = wb[f'KMeans_testing_results']
        ws_hierarchical = wb[f'hierarchical_testing_results']

    column_improvement_tub = "D"
    column_error_tub = "M"
    column_error_basecase = "L"

    if MAEVsRMSE == "MAE" and trainVsTest == "Test":
        if input_period == "year":
            if resolution == "15_min":
                row = 18
            elif resolution == "hourly":
                row = 19
            else:
                row = 20
        elif input_period == "month":
            if resolution == "15_min":
                row = 21
            elif resolution == "hourly":
                row = 22
            else:
                row = 23
        elif input_period == "seasonal":
            if resolution == "15_min":
                row = 24
            elif resolution == "hourly":
                row = 25
            else:
                row = 26
    elif MAEVsRMSE == "MAE" and trainVsTest == "Train":
        if input_period == "year":
            if resolution == "15_min":
                row = 5
            elif resolution == "hourly":
                row = 6
            else:
                row = 7
        elif input_period == "month":
            if resolution == "15_min":
                row = 8
            elif resolution == "hourly":
                row = 9
            else:
                row = 10
        elif input_period == "seasonal":
            if resolution == "15_min":
                row = 11
            elif resolution == "hourly":
                row = 12
            else:
                row = 13
    elif MAEVsRMSE == "RMSE" and trainVsTest == "Test":
        if input_period == "year":
            if resolution == "15_min":
                row = 5
            elif resolution == "hourly":
                row = 6
            else:
                row = 7
        elif input_period == "month":
            if resolution == "15_min":
                row = 8
            elif resolution == "hourly":
                row = 9
            else:
                row = 10
        elif input_period == "seasonal":
            if resolution == "15_min":
                row = 11
            elif resolution == "hourly":
                row = 12
            else:
                row = 13
    else:
        if input_period == "year":
            if resolution == "15_min":
                row = 18
            elif resolution == "hourly":
                row = 19
            else:
                row = 20
        elif input_period == "month":
            if resolution == "15_min":
                row = 21
            elif resolution == "hourly":
                row = 22
            else:
                row = 23
        elif input_period == "seasonal":
            if resolution == "15_min":
                row = 24
            elif resolution == "hourly":
                row = 25
            else:
                row = 26

    for row in range(5, 14):
        ws_KMeans[f"{column_error_basecase}{row}"].value = error_basecase
        ws_hierarchical[f"{column_error_basecase}{row}"].value = error_basecase
    for row in range(18,27):
        ws_KMeans[f"{column_error_basecase}{row}"].value = error_basecase
        ws_hierarchical[f"{column_error_basecase}{row}"].value = error_basecase

    ws_KMeans[f"{column_error_tub}{row}"].value = error_tub
    ws_hierarchical[f"{column_error_tub}{row}"].value = error_tub
    ws_KMeans[f"{column_improvement_tub}{row}"].value = improvement_tub
    ws_hierarchical[f"{column_improvement_tub}{row}"].value = improvement_tub
        
    wb.save(file_path)


def reading_hyperparameters(input_scenario, path_folder, dataset_chosen, clustering_method, resolution, input_period):
    file_path = path_folder / f"Clustering_overview_{dataset_chosen}.xlsx"
    
    wb = load_workbook(file_path, data_only=True)
    ws_hyperparameters = wb['Hyperparameters']

    if input_scenario == "PF_time":
        column_hyperparameters_clusters = "C"
        column_hyperparameters_random_state = "J"
        column_hyperparameters_PCA = "C"
    elif input_scenario == "PQ_Pearson":
        column_hyperparameters_clusters = "D"
        column_hyperparameters_random_state = "K"
        column_hyperparameters_PCA = "D"
    elif input_scenario == "PQSPF_Pearson":
        column_hyperparameters_clusters = "E"
        column_hyperparameters_random_state = "L"
        column_hyperparameters_PCA = "E"
    elif input_scenario == "PQ_PCA":
        column_hyperparameters_clusters = "F"
        column_hyperparameters_random_state = "M"
        column_hyperparameters_PCA = "F"
    elif input_scenario == "PQSPF_PCA":
        column_hyperparameters_clusters = "G"
        column_hyperparameters_random_state = "N"
        column_hyperparameters_PCA = "G"

    if input_period == "year":
        if resolution == "15_min":
            row_KMeans = 5
            row_hierarchical = 18
            row_PCA = 31
        elif resolution == "hourly":
            row_KMeans = 6
            row_hierarchical = 19
            row_PCA = 32
        else:
            row_KMeans = 7
            row_hierarchical = 20
            row_PCA = 33
    elif input_period == "month":
        if resolution == "15_min":
            row_KMeans = 8
            row_hierarchical = 21
            row_PCA = 34
        elif resolution == "hourly":
            row_KMeans = 9
            row_hierarchical = 22
            row_PCA = 35
        else:
            row_KMeans = 10
            row_hierarchical = 23
            row_PCA = 36
    elif input_period == "seasonal":
        if resolution == "15_min":
            row_KMeans = 11
            row_hierarchical = 24
            row_PCA = 37
        elif resolution == "hourly":
            row_KMeans = 12
            row_hierarchical = 25
            row_PCA = 38
        else:
            row_KMeans = 13
            row_hierarchical = 26
            row_PCA = 39
    
    if clustering_method == "KMeans":
        Num_clusters_raw = ws_hyperparameters[f"{column_hyperparameters_clusters}{row_KMeans}"].value
        Num_clusters_list = str(Num_clusters_raw).split(", ")
        Num_clusters = [int(x) for x in Num_clusters_list]
        random_state_raw = ws_hyperparameters[f"{column_hyperparameters_random_state}{row_KMeans}"].value
        random_state_list = str(random_state_raw).split(", ")
        random_state = [int(x) for x in random_state_list]
    else:
        Num_clusters_raw = ws_hyperparameters[f"{column_hyperparameters_clusters}{row_hierarchical}"].value
        Num_clusters_list = str(Num_clusters_raw).split(", ")
        Num_clusters = [int(x) for x in Num_clusters_list]
        random_state = [0]

    if input_scenario in ["PQ_PCA", "PQSPF_PCA"]:
        Num_PCA_components_raw = ws_hyperparameters[f"{column_hyperparameters_PCA}{row_PCA}"].value
        Num_PCA_components_list = str(Num_PCA_components_raw).split(", ")
        Num_PCA_components = [int(x) for x in Num_PCA_components_list]
        correlating_features_list = [0]
    elif input_scenario in ["PQ_Pearson", "PQSPF_Pearson"]:
        correlating_features_raw = ws_hyperparameters[f"{column_hyperparameters_PCA}{row_PCA}"].value
        if input_period == "seasonal":
            correlating_features_list = str(correlating_features_raw).split(" / ")
        else:
            correlating_features_list = str(correlating_features_raw).split(", ")
        Num_PCA_components = [0]
    else:
        Num_PCA_components = [0]
        correlating_features_list = [0]
  
    return Num_clusters, random_state, Num_PCA_components, correlating_features_list

def grouping_error(label_houses, best_PFs, houses, dataset_chosen):
    """
    This function groups the houses based on their labels and returns a list of tuples containing each house and its corresponding best performance factor (PF).
    
    Args:
        label_houses (list): A list of lists, where each inner list contains the houses belonging to a specific cluster.
        best_PFs (dict): A dictionary containing the best performance factors for each group.
        houses (list): A list of all houses.

    Returns:
        PF_list (list): A list of tuples, where each tuple contains a house and its corresponding best PF.
        groups (list): A list of lists, where each inner list contains the houses belonging to a specific group.
    """
    if dataset_chosen == "Slovakia":
        minimal_cluster_size = 15
    else:
        minimal_cluster_size = 1

    sorted_clusters = []
    for i in range(len(label_houses)):
        sorted_clusters.append((i, len(label_houses[i])))
    sorted_clusters.sort(reverse =True, key=lambda x: x[1])
    #print(sorted_clusters)

    groups = [[] for i in range(len(sorted_clusters))]
    rest_group = []

    for i, (group, size) in enumerate(sorted_clusters):
        if size > minimal_cluster_size:
            for item in label_houses[group]:
                groups[i].append(item)
        else:
            for item in label_houses[group]:
                rest_group.append(item)
    #print(groups)
    #print(rest_group)
    key = 1
    house_to_group = {}
    for i, group in enumerate(groups):
        if len(group) > 0:
            group_name = f"group{key}"
            key += 1
            for house in group:
                house_to_group[house] = group_name
    #print(house_to_group)

    best_PFs1 = best_PFs.copy()
    if 'rest' in best_PFs1 and len(rest_group) > 0:
        best_PFs1[f"group{key}"] = best_PFs1.pop('rest')
        for house in rest_group:
            house_to_group[house] = f"group{key}"
    #print(house_to_group)
    #print(best_PFs1)
    PF_list = []
    for house in houses:
        group_name = house_to_group[house]
        PF_list.append((house, best_PFs1[group_name]))

    return(PF_list, groups)


def cluster_hier_decision(data_array, dataset_chosen):
    """Determines the optimal number of clusters for hierarchical clustering using silhouette score and inertia.

    Args:
        data_array (np.ndarray): The input data array to cluster.
        dataset_chosen (str): The name of the dataset chosen.

    Returns:
        tuple: A tuple containing the silhouette scores and inertia values for each number of clusters.
    """
    if dataset_chosen == "Slovakia":
        cluster_range_begin = 2
        cluster_range_end = 30
        select_range_begin = 4
        select_range_end = 11
    else:
        cluster_range_begin = 2
        cluster_range_end = 10
        select_range_begin = 2
        select_range_end = 10


    condensed_dist_matrix = pdist(data_array, metric='euclidean')
    linkage_matrix = linkage(condensed_dist_matrix, method='ward')

    plt.figure(figsize=(10, 5))
    dendrogram(linkage_matrix, leaf_rotation=90)
    plt.title("Hierarchical Clustering Dendrogram (Euclidean Distance)")
    plt.xlabel("Household")
    plt.ylabel("Distance")
    plt.show()

    silhouette_var = []
    inertia_var = []
    x_values = range(cluster_range_begin, cluster_range_end)
    for i in x_values:
        clusters = fcluster(linkage_matrix, i, criterion='maxclust')
        inertia = 0
        for cluster in np.unique(clusters):
            cluster_points = data_array[np.where(clusters == cluster)]
            centroid = np.mean(cluster_points, axis=0)
            inertia += np.sum((cluster_points - centroid) ** 2)/len(data_array)
        inertia_var.append(inertia)
        silhouette_var.append(silhouette_score(squareform(condensed_dist_matrix), clusters, metric='precomputed'))
    
    inertia_scaled = np.array(inertia_var).max() - np.array(inertia_var)  
    inertia_scaled = MinMaxScaler().fit_transform(inertia_scaled.reshape(-1, 1)).flatten()
    silhouette_scaled = MinMaxScaler().fit_transform(np.array(silhouette_var).reshape(-1, 1)).flatten()
    weighted_scores = 0.60 * silhouette_scaled + 0.40 * inertia_scaled
    filtered_scores = [weighted_scores[x_values.index(z)] for z in range(select_range_begin, select_range_end)]
    optimal_k = x_values[np.argmax(filtered_scores)]
    print(optimal_k, silhouette_var[optimal_k-2], inertia_var[optimal_k-2])

    knee_locator = KneeLocator(x_values, inertia_var, curve="convex", direction="decreasing")
    elbow_k = knee_locator.elbow

    fig, ax = plt.subplots(1, 2, figsize=(15, 5))
    ax[0].plot(x_values, inertia_var, marker='x', color='b')
    if elbow_k != None:
        ax[0].axvline(x=elbow_k, color='g')
    ax[0].grid(True)
    ax[0].set_xlabel('Number of clusters')
    ax[0].set_ylabel('Inertia')
    ax[1].plot(x_values, silhouette_var, marker='x', color='b')
    ax[1].grid(True)
    ax[1].set_xlabel('Number of clusters')
    ax[1].set_ylabel('Silhouette Score')
    plt.show()

    return silhouette_var, inertia_var

def cluster_elbow_decision(data_array, dataset_chosen, Clustering_method):
    """Determines the optimal number of clusters for KMeans clustering using the elbow method and silhouette score.

    Args:
        data_array (np.ndarray): The input data array to cluster.
        dataset_chosen (str): The name of the dataset chosen.
        Clustering_method (str): The clustering method to use.
    Returns:
        tuple: A tuple containing the silhouette scores and inertia values for each number of clusters.
    """
    if dataset_chosen == "Slovakia":
        cluster_range_begin = 2
        cluster_range_end = 20
        select_range_begin = 4
        select_range_end = 11
    else:
        cluster_range_begin = 2
        cluster_range_end = 10
        select_range_begin = 2
        select_range_end = 10

    silhouette = []
    inertia = []
    for k in range(cluster_range_begin, cluster_range_end):
        if Clustering_method == "PF_time":
            clusterer = TimeSeriesKMeans(n_clusters=k, n_init=4, metric = 'euclidean', n_jobs=-1, verbose= False, max_iter=50, random_state = 42)
        else:
            clusterer = KMeans(n_clusters=k, n_init=4, verbose= False, max_iter=50, random_state = 42)
        labels = clusterer.fit(data_array)
        inertia.append(clusterer.inertia_)
        silhouette.append(silhouette_score(data_array, clusterer.labels_))

    x_values = range(cluster_range_begin, cluster_range_end)
    
    knee_locator = KneeLocator(x_values, inertia, curve="convex", direction="decreasing")
    elbow_k = knee_locator.elbow
    inertia_scaled = np.array(inertia).max() - np.array(inertia)  
    inertia_scaled = MinMaxScaler().fit_transform(inertia_scaled.reshape(-1, 1)).flatten()
    silhouette_scaled = MinMaxScaler().fit_transform(np.array(silhouette).reshape(-1, 1)).flatten()
    weighted_scores = 0.60 * silhouette_scaled + 0.40 * inertia_scaled
    filtered_scores = [weighted_scores[x_values.index(z)] for z in range(select_range_begin, select_range_end)]
    if dataset_chosen == "Slovakia":
        optimal_k = x_values[np.argmax(filtered_scores)] + 2
    else:
        optimal_k = x_values[np.argmax(filtered_scores)]
    print(optimal_k, silhouette[optimal_k-2], inertia[optimal_k-2])

    fig, ax = plt.subplots(1, 2, figsize=(15, 5))
    ax[0].plot(x_values, inertia, marker='x', color='b')
    if elbow_k != None:
        ax[0].axvline(x=elbow_k, color='g')
    ax[0].grid(True)
    ax[0].set_xlabel('Number of clusters')
    ax[0].set_ylabel('Inertia')
    ax[1].plot(x_values, silhouette, marker='x', color='b')
    ax[1].grid(True)
    ax[1].set_xlabel('Number of clusters')
    ax[1].set_ylabel('Silhouette Score')
    plt.show()

    return silhouette, inertia

def find_random_state(data_array, Num_cluster, Clustering_method):
    """Finds the optimal random state for KMeans clustering based on the Fowlkes-Mallows score.

    Args:
        data_array (np.ndarray): The input data array to cluster.
        Num_cluster (int): The number of clusters to form.
        Clustering_method (str): The clustering method to use.
    
    Returns:
        int: The optimal random state for KMeans clustering.
    """
    Clusterings =[]

    if Clustering_method == "PF_time":
        m = 50
    elif Clustering_method == "PQ_features":
        m = 100

    for i in range(0,m):
        if Clustering_method == "PF_time":
            algort = TimeSeriesKMeans(n_clusters=Num_cluster, random_state=i).fit(data_array)
        else:
            algort = KMeans(n_clusters=Num_cluster, random_state=i).fit(data_array)
        Clusterings.append(algort.labels_)

    Sim_Matrix = pd.DataFrame(0.0,index= ['Clustering{}'.format(i) for i in range(1,m+1)], 
                            columns=['Clustering{}'.format(i) for i in range(1,m+1)])
    for i, inx in enumerate(Sim_Matrix.index):
        for j, col in enumerate(Sim_Matrix.columns):
            Sim_Matrix.at[inx,col] = fowlkes_mallows_score(Clusterings[i],Clusterings[j])

    rounded_val = Sim_Matrix.iloc[:,0].round(2)
    rounded_val = rounded_val.tolist()
    count = Counter(rounded_val)
    most_common = count.most_common(1)
    first_index = rounded_val.index(most_common[0][0])
    print(first_index, most_common)
    random_STATE = first_index
    return random_STATE

def cluster_kmeans(data_array, Num_cluster, random_STATE, houses_list, Clustering_method):
    """Clusters the input data array using KMeans clustering and returns the labeled houses and silhouette score.

    Args:
        data_array (np.ndarray): The input data array to cluster.
        Num_cluster (int): The number of clusters to form.
        random_STATE (int): The random state for reproducibility.
        houses_list (list): A list of house identifiers corresponding to the data array.
    Returns:
        tuple: A tuple containing the labeled houses and silhouette score.
    """
    if Clustering_method == "PF_time":
        km_function = TimeSeriesKMeans(Num_cluster, n_init=4, metric="euclidean", verbose= False, max_iter=50, random_state = random_STATE)
    else:
        km_function = KMeans(Num_cluster, n_init=4, verbose= False, max_iter=50, random_state = random_STATE)
    labels = km_function.fit_predict(data_array)
    labeled_houses = {}
    silhouette = silhouette_score(data_array, labels)
    print(silhouette)
    for house, label in zip(houses_list, labels):
        label = int(label)
        if label not in labeled_houses:
            labeled_houses[label] = []
        labeled_houses[label].append(house)
    print(labeled_houses)
    return(labeled_houses, silhouette)

def cluster_hierarchical(data_array, Num_cluster, houses_list):
    """Clusters the input data array using hierarchical clustering and returns the labeled houses and silhouette score.
    
    Args:
        data_array (np.ndarray): The input data array to cluster.
        Num_cluster (int): The number of clusters to form.
        houses_list (list): A list of house identifiers corresponding to the data array.
    Returns:
        tuple: A tuple containing the labeled houses and silhouette score.
    """
    condensed_dist_matrix = pdist(data_array, metric='euclidean')
    linkage_matrix = linkage(condensed_dist_matrix, method='ward')
    clusters = fcluster(linkage_matrix, Num_cluster, criterion='maxclust') -1
    silhouette = (silhouette_score(squareform(condensed_dist_matrix), clusters, metric='precomputed'))
    print(silhouette)
    labeled_houses = {}
    for house, label in zip(houses_list, clusters):
        label = int(label)
        if label not in labeled_houses:
            labeled_houses[label] = []
        labeled_houses[label].append(house)
    print(labeled_houses)
    return(labeled_houses, silhouette)

def grouping(label_houses, dataset_chosen):
    """Groups the labeled houses into clusters based on the specified dataset.

    Args:
        label_houses (dict): A dictionary containing labeled houses.
        dataset_chosen (str): The name of the dataset chosen.

    Returns:
        list: A list of groups containing houses.
    """

    if dataset_chosen == "Slovakia":
        minimal_cluster_size = 15
    else:
        minimal_cluster_size = 1

    sorted_clusters = []
    
    for i in range(len(label_houses)):
        sorted_clusters.append((i, len(label_houses[i])))
    sorted_clusters.sort(reverse =True, key=lambda x: x[1])
    print(sorted_clusters)

    groups = [[] for _ in range(len(sorted_clusters))]

    for i, (group, size) in enumerate(sorted_clusters):
        if size > minimal_cluster_size:
            for item in label_houses[group]:
                    groups[i].append(item)
    return(groups)

def error_calc(groups2, time_series, MAEVsRMSE):
    """Calculates the estimation error and improvement compared to basecase made when clustering

    Args:
        groups2 (list): List of groups containing houses.
        time_series (dict): Dictionary containing time series data for each house.
        MAEVsRMSE (str): The error metric used ("MAE" or "RMSE").
    
    Returns:
        tuple: A tuple containing the minimum errors and best power factors for each group.
    """
    house_to_group = {}
    rest = []
    for i, group in enumerate(groups2):
        if len(group) > 0:
            group_name = f"group{i+1}"
            for house in group:
                house_to_group[house] = group_name

    groups = ["rest"] + [f"group{i+1}" for i in range(len(groups2)) if len(groups2[i]) > 0]
    min_errors = {group: float('inf') for group in groups}
    best_PFs = {group: None for group in groups}


    PF_range = [round(x * 0.01, 2) for x in range(61, 101)]

    house_data = {}
    for ind, power_data in time_series.items():
        house = power_data['Nr'].iloc[0]
        house_data[house] = {
            'P': power_data['P'].to_numpy(),
            'Q': power_data['Q'].to_numpy()
        }

    for PF in PF_range:
        current_errors = {group: 0 for group in groups}

        for house, data in house_data.items():
            P = data['P']
            Q = data['Q']
            
            Q_est_inner = np.sqrt((P / PF) ** 2 - P ** 2)
            if MAEVsRMSE == "MAE":
                error = (abs(Q-Q_est_inner)).sum()
            else:
                error = ((Q - Q_est_inner) ** 2).sum()
        
            group = house_to_group.get(house, 'rest')
            if group == 'rest' and house not in rest:
                rest.append(house)
            current_errors[group] += error

        for group in groups:
            if current_errors[group] < min_errors[group]:
                min_errors[group] = current_errors[group]
                best_PFs[group] = PF

    for group in groups:
        if group == "rest":
            group_number = len(rest)
            if group_number > 0:
                print(group_number)
                if MAEVsRMSE == "MAE":
                    print(f"Minimum Error for {group}: {round(min_errors[group]/group_number/len(time_series[1]['P']), 5)}")
                else:
                    print(f"Minimum Error for {group}: {round(np.sqrt(min_errors[group]/group_number/len(time_series[1]['P'])),5)}")
                print(f"Best PF for {group}: {best_PFs[group]}")
        else:
            group_number = int(group[5:])-1
            if len(groups2[group_number]) > 0:
                print(len(groups2[group_number]))
                if MAEVsRMSE == "MAE":
                    print(f"Minimum Error for {group}: {round(min_errors[group]/len(groups2[group_number])/len(time_series[1]['P']),5)}")
                else:
                    print(f"Minimum Error for {group}: {round(np.sqrt(min_errors[group]/len(groups2[group_number])/len(time_series[1]['P'])),5)}")
                print(f"Best PF for {group}: {best_PFs[group]}")

    return(min_errors, best_PFs)
















